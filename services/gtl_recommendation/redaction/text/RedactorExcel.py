from pathlib import Path
import json, os, re, pandas as pd, csv, logging
import uuid
from typing import List, Dict
import openpyxl
 
# Assuming these modules are available in your structure
from core.db.crud import DatabaseManager
from core.exceptions import NoSensitiveItemFound
from core.utility import get_custom_logger
from config import Config as cfg
logger = get_custom_logger(__name__)
# logger.propagate = False
 
class ExcelRedactor:
    # _InclusionDf = pd.read_csv(Path(r"services/process_ips/redaction/helperdata/Inclusion.csv"))
    # _EXCLUSION_ITEMS = pd.read_csv(Path(r"services/process_ips/redaction/helperdata/Exclusion.csv")).Exclude.to_list()
   
    # sanitize_data = []
 
    def __init__(self, filepath: Path, dafileid: uuid = None, debug: bool = False):
        self.filepath = filepath
        self.dafileid = dafileid

        self.records: List[Dict] = []
        self.workbook = openpyxl.load_workbook(filepath)
        self.filecontent: str = ''
        self.sensitive_info: List[Dict]=[]

        if debug:
            logger.setLevel(logging.DEBUG)
 
    def _replace_text_in_cell(self, ws, row_idx, col_idx, cell_value, label, pattern):
        """Performs the sensitive text replacement within a single cell value.
 
        The function now tolerates `label` being a float (e.g. NaN) by coercing it
        to a string before calling `[replace](cci:1://file:///c:/Users/Vedish_Kabara/redaction/ip_redaction/redactors/text/RedactorDoc.py:567:4-619:24)`.
        """
        label = "" if pd.isna(label) else str(label)
        safe_pattern = self.escape_custom(str(pattern))
        # Handle quote variations in the pattern
        # if "'" in safe_pattern:
        if any(quote in safe_pattern for quote in '`\'’'):
            # Replace escaped single quote with a character class that matches both ' and `
            quote_variations = safe_pattern.replace("\\'", "[`'’]?")
        else:
            quote_variations = safe_pattern
        # Add word boundaries to avoid matching within words
        word_boundary_pattern = rf'\b{quote_variations}\b'
        # word_boundary_pattern = rf'(?<![a-zA-Z0-9_]){safe_pattern}(?![a-zA-Z0-9_])'

        len_diff = 0
        # Find and replace every occurrence inside the cell value
        for m in re.finditer(word_boundary_pattern, cell_value, re.IGNORECASE):
            try:
                # Prefer the first capturing group if it exists (same as DocRedactor)
                match_str = m.group(1)
                startIdx = m.start() + m.group().find(m.group(1))
                endIdx = startIdx + len(m.group(1))
                # match_str = m.group(1)
                # startIdx = m.start(1)  # Use m.start(1) directly for group 1 start
                # endIdx = m.end(1)      # Use m.end(1) directly for group 1 end
            except IndexError:
                # Fallback to the whole match
                match_str = m.group(0)
                startIdx = m.start()
                endIdx = m.end()
 
            # Skip anything that is in the exclusion list
            # if match_str not in ExcelRedactor._EXCLUSION_ITEMS:
            placeholder = f"<{cfg.PREFIX}{label.replace(' ', '_')}>"
 
            # Record the redaction for later reporting / DB insert
            self.records.append({
                "start": startIdx,
                "end": endIdx,
                "label": label,
                "sensitivetext": match_str,
                "placeholder": placeholder,
                "context": (
                    f"Sheet: {ws.title}, Cell: {col_idx}{row_idx}, Text: "
                    + cell_value[max(0, startIdx - 40):min(len(cell_value), endIdx + 40)]
                )
            })
 
            # Apply the placeholder to the cell value
            cell_value = cell_value[:startIdx+len_diff] + placeholder + cell_value[endIdx+len_diff:]
            #calculate difference in len of string for next iteration
            len_diff = len_diff + len(placeholder)-(endIdx-startIdx)
        
        if cell_value.startswith('<') and cell_value.endswith('>'):
            cell_value = ''
        #  Return the (possibly) modified cell value
        return cell_value
 
    def _redact_excel(self, label, pattern):
        """Iterates over all sheets and cells to apply redaction."""
        for sheetname in self.workbook.sheetnames:
            ws = self.workbook[sheetname]
           
            # Use iter_rows to read and set cell values
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    cell_value = cell.value
                    if cell_value is not None and str(cell_value).strip() != '':
                        original_value = str(cell_value)
                        
                        # Call the replacement logic
                        new_value = self._replace_text_in_cell(
                            ws, row_idx, openpyxl.utils.get_column_letter(col_idx), original_value, label, pattern
                        )
                        
                        # Update the cell's value only if it was modified
                        if new_value != original_value:
                            print('orig-', original_value)
                            print('changed-', new_value)
                             # Convert the column index back to letter for the cell reference
                            cell_ref = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                            # ws[cell_ref] = new_value
                           
                            # Copy all formatting properties from the original cell
                            old_cell = ws[cell_ref]

                            # Set the new value and preserve all formatting
                            ws[cell_ref] = new_value
                            if old_cell.font:
                                from openpyxl.styles import Font
                                new_font = Font(
                                    name=old_cell.font.name,
                                    size=old_cell.font.size,
                                    bold=old_cell.font.bold,
                                    italic=old_cell.font.italic,
                                    underline=old_cell.font.underline,
                                    strike=old_cell.font.strike,
                                    color=old_cell.font.color
                                )
                                ws[cell_ref].font = new_font
                            if old_cell.fill:
                                from openpyxl.styles import PatternFill, GradientFill
                                if hasattr(old_cell.fill, 'patternType'):
                                    # PatternFill
                                    new_fill = PatternFill(
                                        patternType=old_cell.fill.patternType,
                                        fgColor=old_cell.fill.fgColor,
                                        bgColor=old_cell.fill.bgColor
                                    )
                                    ws[cell_ref].fill = new_fill
                                else:
                                    # GradientFill or other fill types
                                    ws[cell_ref].fill = old_cell.fill
                            if old_cell.border:
                                from openpyxl.styles import Border
                                new_border = Border(
                                    left=old_cell.border.left,
                                    right=old_cell.border.right,
                                    top=old_cell.border.top,
                                    bottom=old_cell.border.bottom,
                                    diagonal=old_cell.border.diagonal,
                                    diagonal_direction=old_cell.border.diagonal_direction,
                                    outline=old_cell.border.outline,
                                    start=old_cell.border.start,
                                    end=old_cell.border.end
                                )
                                ws[cell_ref].border = new_border
                            if old_cell.alignment:
                                from openpyxl.styles import Alignment
                                new_alignment = Alignment(
                                    horizontal=old_cell.alignment.horizontal,
                                    vertical=old_cell.alignment.vertical,
                                    text_rotation=old_cell.alignment.text_rotation,
                                    wrap_text=old_cell.alignment.wrap_text,
                                    shrink_to_fit=old_cell.alignment.shrink_to_fit,
                                    indent=old_cell.alignment.indent
                                )
                                ws[cell_ref].alignment = new_alignment
                            if old_cell.number_format:
                                ws[cell_ref].number_format = old_cell.number_format
                        
   
    def escape_custom(self, s: str) -> str:
        SPECIAL_CHARS = r'.^$*+?{}[]\|()#\''
        escaped = []
        for ch in s:
            if ch in SPECIAL_CHARS:
                escaped.append('\\' + ch)
            else:
                escaped.append(ch)
        return ''.join(escaped)

    
    def set_sensitiveinfo(self):
        logger.info(f"{self.dafileid}-->Fetching sensitive data from DB")
        db = DatabaseManager()
        df = db.get_vwtoberedacted(requestid=self.requestid, fuuid=self.fuuid, dafileid = self.dafileid)
        if df.empty:
            raise NoSensitiveItemFound()

        df = df[['category', 'sensitivetext']].drop_duplicates()
        df.sort_values(by="sensitivetext", key=lambda x: x.str.len(), inplace=True,ascending=False)        
        self.sensitive_data = df.to_dict(orient='records')
            
    def sanitize(self, **kwargs):    
        self.requestid = kwargs.get('requestid')
        self.fuuid = kwargs.get('fuuid')
        self.dafileid = kwargs.get('dafileid')
        
        self.set_sensitiveinfo()  

        for tr in self.sensitive_data:
            self._redact_excel(label=tr['category'], pattern=tr['sensitivetext'])
        
        totalRedacted = len(self.records)
        return len(self.records) > 0, totalRedacted
   
    # # Keeping saveJson for debug logs (optional)
    def saveJson(self, outdir: str, jsondata, postfix: str = ""):
        try:
            json_path = os.path.join(outdir, f"{self.filepath.stem}{postfix}.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(jsondata, f, indent=2)
        except Exception as e:
            logger.error(f"Saving JSON failed at '{json_path}': {e}", exc_info=True)
 
    def save(self, outdir: str, filename:str = None):
        """Saves the redacted workbook and the redaction records CSV."""
        try:
            # 2. Construct the full output path
            fname = filename if filename is not None else self.filepath.name
            OutPathFile = Path(os.path.join(outdir,fname))
            OutPathFile.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(OutPathFile)
            logger.info(f"Saved redacted xlsx: {OutPathFile}")

            if len(self.records)>0:
                db = DatabaseManager()
                db.insert_redacted_results(self.requestid, self.fuuid, self.dafileid, fname, self.records)
                logger.info(f"Redacted {len(self.records)} sensitive items")
                
                redactedItemsFilepath = os.path.join(outdir, f"{OutPathFile.stem}_redacted.csv")
                # Save the redaction records CSV (using the original stem for the CSV name)
                pd.DataFrame(pd.json_normalize(self.records)).drop(['start', 'end'], axis=1, errors='ignore').to_csv(
                    redactedItemsFilepath,
                    index=False,
                    quoting=csv.QUOTE_ALL
                )
                return Path(OutPathFile),Path(redactedItemsFilepath)
            else:
                return Path(OutPathFile),None
        except Exception as e:
            logger.error(f"Saving XLSX failed at '{OutPathFile}': {e}", exc_info=True)
            raise
        finally:
            self.workbook.close()